from pprint import pprint
import json
import openpyxl
# import pdftables_api
from pathlib import Path
from utilproject.settings import PDF_TABLE_API_KEY


def __init__():
    return


def parse_xlsx(path):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[workbook.sheetnames[0]]

    bareme = {}
    localisations = []
    """ determination de la derniere cellule utilisée ex G148 """
    dim = sheet.calculate_dimension().split(':')[-1]

    """ recuperation de la partie numérique ex:148 = ligne max """
    res = ''.join(i for i in dim if i.isdigit())
    max_l = int(res)

    for col in sheet.iter_cols(max_col=1, min_row=4):
        departements = list({cell.value[:2] for cell in col})
        print(departements)

        for departement in departements:
            for row in sheet.iter_rows(min_row=4, max_col=1):
                for cell in row:
                    if cell.value[:2] == departement:
                        localisations.append(cell.value)

                        for localisation in localisations:
                            bareme[localisation] = {"M": {"R": 0, "N+PD": 0},
                                                    "C": {"R": 0, "N+PD": 0},
                                                    "ACOSS": {"R": 0, "N+PD": 0}
                                                    }
            # localisations = []

    for i in range(4, max_l + 1):
        loc = sheet.cell(i, 1).value
        acoss_r = sheet.cell(i, 6).value
        acoss_npd = sheet.cell(i, 7).value
        cadre_r = sheet.cell(i, 4).value
        cadre_npd = sheet.cell(i, 5).value
        noncadre_npd = sheet.cell(i, 3).value
        noncadre_r = sheet.cell(i, 2).value

        bareme[loc]["M"]["R"] = noncadre_r
        bareme[loc]["C"]["R"] = cadre_r
        bareme[loc]["ACOSS"]["R"] = acoss_r
        bareme[loc]["M"]["N+PD"] = noncadre_npd
        bareme[loc]["C"]["N+PD"] = cadre_npd
        bareme[loc]["ACOSS"]["N+PD"] = acoss_npd

    workbook.close()

    return bareme


def save_to_json(datas, name=""):
    with open(name, "w") as f:
        json.dump(datas, f, indent=4)
    return


def get_json(name):
    try:
        with open(name, "r") as f:
            fichier = json.load(f)
    except:
        return False
    return fichier


def update_ursaff(annee, taux_cs_ecart, taux_cs_non_soumises):
    ursaff = get_json('../static/datas/ursaff.json')
    if not ursaff:
        ursaff = {annee: {'taux_cs_ecart': taux_cs_ecart,
                          'taux_cs_non_soumises': taux_cs_non_soumises,
                          }
                  }

    else:
        ursaff[annee] = {'taux_cs_ecart': taux_cs_ecart,
                         'taux_cs_non_soumises': taux_cs_non_soumises,
                         }
    save_to_json(ursaff, name='../static/datas/ursaff.json')


def jsonTodb(file):

    with open(file, 'r') as f:
        data = json.load(f)

    for element in data.keys():
        for x in data[element].keys():
            print(x)
            repas = data[element][x]['R']
            nuit = data[element][x]['N+PD']

            print(element, x, repas, nuit)


def xlsx_to_db(path, an):

    datas = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[workbook.sheetnames[0]]

    """ determination de la derniere cellule utilisée ex A1:G148 """
    dim = sheet.calculate_dimension().split(':')[-1]

    """ recuperation de la partie numérique ex:148 = ligne max """
    res = ''.join(i for i in dim if i.isdigit())
    max_l = int(res)

    for i in range(4, max_l + 1):

        data1 = {}
        data2 = {}
        data3 = {}

        data1["annee"] = an
        data1["localisation"] = sheet.cell(i, 1).value
        data1["college"] = "M"
        data1["Nuit_Pdj"] = sheet.cell(i, 3).value
        data1["Repas"] = sheet.cell(i, 2).value
        datas.append(data1)

        data2["annee"] = an
        data2["localisation"] = sheet.cell(i, 1).value
        data2["college"] = "C"
        data2["Nuit_Pdj"] = sheet.cell(i, 5).value
        data2["Repas"] = sheet.cell(i, 4).value
        datas.append(data2)

        data3["annee"] = an
        data3["localisation"] = sheet.cell(i, 1).value
        data3["college"] = "ACOSS"
        data3["Nuit_Pdj"] = sheet.cell(i, 7).value
        data3["Repas"] = sheet.cell(i, 6).value
        datas.append(data3)

    return datas


def pdf_to_xlsx(path, output='.'):
    # c = pdftables_api.Client(PDF_TABLE_API_KEY)
    # c.xlsx(path, output)
    if Path.is_file(output):
        success = True
    else:
        success = False

    return success


def xlsx_to_dict(path):

    workbook = openpyxl.load_workbook(path)
    sheets = workbook.worksheets
    annee = sheets[0]['A1'].value[-4:]

    datas = []

    for sheet in sheets:
        datas1 = {}
        dim = sheet.calculate_dimension().split(':')[-1]
        """ recuperation de la partie numérique ex:148 = ligne max """
        res = ''.join(i for i in dim if i.isdigit())
        max_l = int(res)
        # max_c = 9
        if sheet['A1'].value:
            annee = sheet['A1'].value[-4:]
            maj_res = float(sheet['G1'].value.split(':')[-1].replace(',', '.'))
            debut = 4

        else:
            debut = 2

            if sheet.title == 'Page 3':
                return

        for i in range(debut, max_l + 1):
            datas1 = {}
            if isinstance(sheet.cell(i, 1).value, int):

                datas1['NR'] = sheet.cell(i, 1).value
                datas1['coeff'] = float(sheet.cell(i, 2).value.replace(',', '.').replace(' ', ''))

            datas.append(datas1)

    return datas


def populate_nr():
    nr_choix = ["30", "35", "40", "45", "50", "55", "60", "65", "70", "75", "80", "85", "90", "95", "100",
                "105", "110", "115", "120", "125", "130", "135", "140", "145", "150", "155", "160", "165",
                "170", "175", "180", "185", "190", "195", "200", "205", "210", "215", "220", "225", "230",
                "235", "240", "245", "250", "255", "260", "265", "270", "275", "280", "285", "290", "295", "300",
                "305", "310", "315", "320", "325", "330", "340", "350", "355", "360", "365", "370"]

    coeff = ["226,0", "230,4", "234,9", "239,6", "244,3", "249,0", "253,8", "259,0", "264,4", "269,4", "274,2",
             "280,5", "286,9", "293,2", "299,8", "306,7", "313,9", "321,7", "330,6", "338,9", "347,1", "355,7",
             "364,5", "373,6", "382,8", "392,1", "403,0", "412,7", "422,8", "433,3", "444,0", "454,9", "466,1",
             "477,6", "489,3", "501,5", "513,9", "526,5", "539,4", "552,9", "566,5", "580,6", "598,5", "613,2",
             "628,3", "643,9", "659,8", "676,2", "692,9", "709,9", "727,5", "744,0", "760,7", "777,6", "794,9",
             "812,6", "830,7", "849,3", "868,5", "887,4", "906,7", "929,0", "949,6", "971,4", "993,8", "1016,7",
             "1040,0"]

    choix = [_.replace(',', '.') for _ in nr_choix]
    coeff = [_.replace(',', '.') for _ in coeff]

    return [{'NR': choix[x], 'valeur': coeff[x]} for x in range(len(coeff))]






if __name__ == '__main__':
    # jsonTodb('/Users/alainzypinoglou/PycharmProject/django-util-replica/static/datas/2022.json')
    # annee = get_json('../static/datas/ursaff.json')
    # bareme = parse_xlsx(path='../static/datas/frais2021.xlsx')
    # xlsx_to_db(path='/Users/alainzypinoglou/PycharmProject/django-util-replica/test_bareme.xlsx', an=2023)
    # save_to_json(datas=bareme, name='2021.json')@
    # pprint(annee)
    # print(Path.cwd())
    # update_ursaff('2021', 8.86, 6.01)
    # pdf_to_xlsx('/Users/alainzypinoglou/Documents/CFDT/snb.pdf', 'test')
    # print(xlsx_to_dict('output.xlsx'))
    populate_nr()


